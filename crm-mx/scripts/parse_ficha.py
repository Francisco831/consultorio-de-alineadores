#!/usr/bin/env python3
"""Parsea la ficha de entregas (data/ficha_entrega.xlsx, tabs 2023-2026 +
SkyDropx/Lalamove/DHL) → data/ficha_tipos.json con, por ID externo:

  tipo_caso: Fast | Medium | Full | Kids | Teens   (derivado de la col Etapa)
  etapas: códigos de entrega vistos, alineadores_total, entregas, ultima_entrega

Nomenclatura col Etapa: 1.0/2.0/... = etapas de un FULL · M = Medium · F = Fast
· K1-K4 = Kids · T1-T2 = Teens · nBis = bis de etapa n · '-' = sin dato.
"""
import json
import re
from datetime import datetime
from pathlib import Path

import openpyxl

BASE = Path(__file__).resolve().parent.parent
OUT = BASE / "data" / "ficha_tipos.json"

TABS = ["2023", "2024", "2025", "2026", "SkyDropx", "Lalamove", "DHL"]


def classify(codes):  # python 3.9: sin sintaxis "str | None"
    up = {c.upper() for c in codes}
    if any(c.startswith("K") for c in up):
        return "Kids"
    if any(c.startswith("T") for c in up):
        return "Teens"
    if "M" in up:
        return "Medium"
    if "F" in up:
        return "Fast"
    if any(re.match(r"^\d", c) for c in up):
        return "Full"
    return None


def main():
    wb = openpyxl.load_workbook(
        BASE / "data" / "ficha_entrega.xlsx", read_only=True, data_only=True
    )
    per_id = {}
    for tab in TABS:
        if tab not in wb.sheetnames:
            continue
        ws = wb[tab]
        header = [str(c or "").strip().lower() for c in next(ws.iter_rows(max_row=1, values_only=True))]

        def col(*names):
            for n in names:
                for i, h in enumerate(header):
                    if n in h:
                        return i
            return None

        i_fecha, i_etapa, i_id, i_alin = (
            col("fecha"),
            col("etapa"),
            col("id"),
            col("alineador"),
        )
        if i_id is None or i_etapa is None:
            continue
        for r in ws.iter_rows(min_row=2, values_only=True):
            if not r or len(r) <= max(i_id, i_etapa):
                continue
            rid = str(r[i_id]).strip().upper() if r[i_id] else None
            if not rid or not re.match(r"^[A-Z]{2}\d+", rid):
                continue
            code = str(r[i_etapa]).strip() if r[i_etapa] is not None else ""
            entry = per_id.setdefault(
                rid, {"codes": set(), "alineadores": 0, "entregas": 0, "ultima": None}
            )
            if code and code != "-":
                entry["codes"].add(code)
            entry["entregas"] += 1
            if i_alin is not None and r[i_alin] is not None:
                try:
                    entry["alineadores"] += int(float(r[i_alin]))
                except (TypeError, ValueError):
                    pass
            f = r[i_fecha] if i_fecha is not None else None
            if isinstance(f, datetime):
                iso = f.date().isoformat()
                if not entry["ultima"] or iso > entry["ultima"]:
                    entry["ultima"] = iso

    out = []
    tipos_count = {}
    for rid, e in per_id.items():
        tipo = classify(e["codes"])
        tipos_count[tipo or "—"] = tipos_count.get(tipo or "—", 0) + 1
        out.append(
            {
                "id_externo": rid,
                "tipo_caso": tipo,
                "etapas": sorted(e["codes"]),
                "alineadores_total": e["alineadores"] or None,
                "entregas": e["entregas"],
                "ultima_entrega": e["ultima"],
            }
        )
    OUT.write_text(json.dumps(out, ensure_ascii=False, indent=1))
    print(f"IDs en ficha: {len(out)} | tipos: {tipos_count}")


if __name__ == "__main__":
    main()
