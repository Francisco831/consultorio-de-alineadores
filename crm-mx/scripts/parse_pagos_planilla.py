#!/usr/bin/env python3
"""Parsea la hoja "Facturación y Cobranzas" de Administración México a pagos.

MISMA lógica y MISMAS claves que parse_enrichment.parse_payments
(adminmx:{fila}:{slot 1-5}) — así el ledger del CRM y finanzas nunca duplican.
Sin DoctorBase: noloco_id sale null y el vínculo doctor lo hace el CRM después.

Fuentes (una de las dos):
  --xlsx PATH   export manual de la sheet (Archivo → Descargar → .xlsx)
  --json PATH   respuesta del Apps Script (gas-pagos-planilla.gs), bajada p.ej.
                con curl "$PLANILLA_MX_URL?secret=$PLANILLA_MX_SECRET" -o /tmp/p.json

Salida: --out (default data/pagos_planilla.json)
"""
import argparse
import json
import os
import re
import sys

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
from parse_enrichment import cell_str, norm_method, to_amount, to_date  # noqa: E402

TAB = "Facturación y Cobranzas"


def cargar_grilla(args):
    if args.xlsx:
        from openpyxl import load_workbook
        wb = load_workbook(args.xlsx, read_only=True, data_only=True)
        if TAB not in wb.sheetnames:
            sys.exit(f"el archivo no tiene la hoja {TAB!r} (hojas: {wb.sheetnames}) — ¿es el export correcto?")
        return [list(r) for r in wb[TAB].iter_rows(values_only=True)]
    with open(args.json) as f:
        data = json.load(f)
    if data.get("tab") != TAB:
        sys.exit(f"el JSON no es de la hoja {TAB!r}: {str(data)[:120]}")
    # las fechas pueden venir como timestamp ISO UTC ("2022-04-22T03:00:00.000Z",
    # medianoche GMT-3/-6): el día calendario es el del string, se corta directo
    iso_ts = re.compile(r"^\d{4}-\d{2}-\d{2}T\d{2}:")
    return [
        [c[:10] if isinstance(c, str) and iso_ts.match(c) else c for c in row]
        for row in data["values"]
    ]


def parse_payments(rows):
    header = [cell_str(c) for c in rows[0]]

    def idx_all(pat):
        return [i for i, c in enumerate(header) if re.fullmatch(pat, c.strip(), re.I)]

    id_cols = [i for i, c in enumerate(header) if c.strip().upper() == "ID"]
    c_id = id_cols[1] if len(id_cols) > 1 else id_cols[0]
    c_pac = next(i for i, c in enumerate(header) if c.strip().upper() == "PACIENTE")
    c_prof = next(i for i, c in enumerate(header) if c.strip().upper() == "PROFESIONAL")
    formas = idx_all(r"FORMA DE PAGO")
    fechas_pago = idx_all(r"FECHA PAGO")
    amounts = [i for i, c in enumerate(header) if re.fullmatch(r"\d\s*°\s*PAGO", c.strip(), re.I)]
    nfacs = idx_all(r"N°\s*FAC")
    assert len(formas) == len(fechas_pago) == len(amounts) == 5, \
        f"la hoja cambió de estructura: formas={formas} fechas={fechas_pago} montos={amounts}"

    payments, undated = [], 0
    for rix, row in enumerate(rows[1:], start=2):  # numeración de fila de la hoja
        cells = list(row) + [None] * max(0, len(header) - len(row))
        case_id = cell_str(cells[c_id])
        if not case_id or case_id.upper() == "ID":
            continue
        prof_raw = cell_str(cells[c_prof])
        for k in range(5):
            amt = to_amount(cells[amounts[k]])
            pdate = to_date(cells[fechas_pago[k]])
            if amt is None or amt == 0:
                continue
            if not pdate:
                undated += 1
                continue
            nfac = cell_str(cells[nfacs[k]]) if k < len(nfacs) and nfacs[k] < len(cells) else ""
            payments.append({
                "external_key": "adminmx:%d:%d" % (rix, k + 1),
                "doctor_nombre_raw": prof_raw or None,
                "noloco_id": None,
                "case_external_id": case_id,
                "paciente": cell_str(cells[c_pac]) or None,
                "amount_mxn": round(amt, 2),
                "paid_at": pdate,
                "method": norm_method(cells[formas[k]]),
                "notes": ("fac:%s" % nfac) if nfac and nfac != "-" else None,
            })
    return payments, undated


def parse_casos(rows):
    """Una fila = un caso(-etapa) con su precio pactado. VALOR (lista menos
    descuentos) es el monto fijo del caso; 0 = caso sin precio o bonificado."""
    header = [cell_str(c) for c in rows[0]]
    def col(nombre):
        return next(i for i, c in enumerate(header) if c.strip().upper() == nombre)
    c_id2, c_tipo, c_etapa = col("ID"), col("TIPO"), col("ETAPA")
    id_cols = [i for i, c in enumerate(header) if c.strip().upper() == "ID"]
    c_id = id_cols[1] if len(id_cols) > 1 else id_cols[0]
    c_envio = next(i for i, c in enumerate(header) if "ENV" in c.upper())
    c_pac, c_prof = col("PACIENTE"), col("PROFESIONAL")
    c_dc = next(i for i, c in enumerate(header) if "D/C" in c.upper())
    c_imp = col("IMPORTE")
    c_cat = next(i for i, c in enumerate(header) if c.strip().upper() == "CATEGORÍA" or c.strip().upper() == "CATEGORIA")
    c_valor = col("VALOR")
    pagos_idx = [i for i, c in enumerate(header) if re.fullmatch(r"\d\s*°\s*PAGO", c.strip(), re.I)]
    casos = []
    for rix, row in enumerate(rows[1:], start=2):
        cells = list(row) + [None] * max(0, len(header) - len(row))
        case_id = cell_str(cells[c_id])
        if not case_id or case_id.upper() == "ID":
            continue
        casos.append({
            "fila": rix,
            "case_id": case_id,
            "tipo": cell_str(cells[c_tipo]) or None,
            "etapa": cell_str(cells[c_etapa]) or None,
            "paciente": cell_str(cells[c_pac]) or None,
            "profesional": cell_str(cells[c_prof]) or None,
            "fecha": to_date(cells[c_dc]) or to_date(cells[c_envio]),
            "categoria": cell_str(cells[c_cat]) or None,
            "importe_lista": to_amount(cells[c_imp]),
            "valor": to_amount(cells[c_valor]) or 0,
            # suma de los 5 slots CON o SIN fecha: es el cobrado real del caso
            # (los movements de finanzas solo arrancan en 2026)
            "pagado": round(sum(to_amount(cells[i]) or 0 for i in pagos_idx), 2),
        })
    return casos


def main():
    ap = argparse.ArgumentParser()
    src = ap.add_mutually_exclusive_group(required=True)
    src.add_argument("--xlsx")
    src.add_argument("--json")
    ap.add_argument("--out", default=os.path.join(os.path.dirname(__file__), "../data/pagos_planilla.json"))
    ap.add_argument("--out-casos", default=os.path.join(os.path.dirname(__file__), "../data/casos_planilla.json"))
    args = ap.parse_args()

    rows = cargar_grilla(args)
    payments, undated = parse_payments(rows)
    with open(args.out, "w") as f:
        json.dump(payments, f, ensure_ascii=False, indent=1)
    casos = parse_casos(rows)
    with open(args.out_casos, "w") as f:
        json.dump(casos, f, ensure_ascii=False, indent=1)
    print(f"{len(casos)} casos -> {os.path.normpath(args.out_casos)}")

    por_mes = {}
    for p in payments:
        if p["paid_at"] >= "2026-01-01":
            k = p["paid_at"][:7]
            por_mes[k] = por_mes.get(k, 0) + p["amount_mxn"]
    print(f"{len(payments)} pagos ({undated} montos sin fecha, salteados) → {os.path.normpath(args.out)}")
    for k in sorted(por_mes):
        print(f"  {k}: {por_mes[k]:,.2f}")


if __name__ == "__main__":
    main()
