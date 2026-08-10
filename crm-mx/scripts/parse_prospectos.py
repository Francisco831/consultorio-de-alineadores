#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
parse_prospectos.py — extrae TODAS las personas de las fuentes de prospección
(universo A: doctores que dieron su contacto y no están en Noloco/plataformas).

A diferencia de parse_enrichment.py (que solo guardaba matches con Noloco),
acá se emite TODO el mundo, con su evento/fuente, para cargarlo al CRM.

Fuentes:
  - madre "Mexico Pancho_Juan": 28 tabs de cursos/meetings/AMO/gira (eventos)
    + tabs Aliwell y Ortodoncistas Cryst (targets con competidor)
  - "Mexico - Drs. INVISALIGN Y CLINICAS": doctores Invisalign por estado,
    con categoría Align y casos/año → estimated_cases_month
  - "Acreditados Zona Norte": Potenciales Align / Cryst aligner / A
  - "B-Learning Bono": Inscriptos (pagaron el curso) y Contactados

Explícitamente EXCLUIDO (directorio masivo, no leads):
  BASE DE DATOS COMPLETA (9.5k), BASE NO ACREDITADOS (7k),
  DRES/BASE CONTACTADOS (5.9k/2.5k), Ortodoncistas no acreditados (1.6k/1.8k).

Output: crm-mx/data/prospectos_fuentes.json
Python 3.9 / openpyxl.
"""
import json
import os
import re
import unicodedata
from datetime import datetime, date

import openpyxl

SCRATCH = ("/private/tmp/claude-501/-Users-franciscobasilico-Desktop-Claude-Periskope/"
           "dfb9ce39-9fac-4ca6-a9b4-daf01a381467/scratchpad/crm_sheets")
DATA = "/Users/franciscobasilico/Desktop/Claude/Periskope/crm-mx/data"

F_MADRE = os.path.join(DATA, "madre_gestion.xlsx")  # export fresco del 8/8
F_NORTE = os.path.join(DATA, "acredi_norte.xlsx")
F_BLEARN = os.path.join(DATA, "blearning_bono.xlsx")
F_INVIS = os.path.join(DATA, "drs_invisalign.xlsx")

# ---------------------------------------------------------------- helpers
TITLE_TOKENS = {"dr", "dra", "drs", "dres", "doctor", "doctora", "cks", "imed", "od",
                "cd", "mtro", "mtra", "esp", "lic"}
PARTICLES = {"de", "del", "la", "las", "los", "y", "e", "da", "do", "van", "der"}

def strip_accents(s):
    return "".join(c for c in unicodedata.normalize("NFD", s)
                   if unicodedata.category(c) != "Mn")

def name_tokens(raw):
    if raw is None:
        return frozenset()
    s = strip_accents(str(raw)).lower()
    s = re.sub(r"\(.*?\)", " ", s)
    s = re.sub(r"^\s*\d+[.)]?\s*", " ", s)   # "1. Dr. Jose ..." de listas numeradas
    s = re.sub(r"[^a-z\s]", " ", s)
    toks = [t for t in s.split() if t not in TITLE_TOKENS and t not in PARTICLES and len(t) > 1]
    return frozenset(toks)

def clean_name(raw):
    """Nombre presentable: saca numeración y títulos, colapsa espacios."""
    s = str(raw).strip()
    s = re.sub(r"^\s*\d+[.)]?\s*", "", s)
    s = re.sub(r"^(dra?\.?|doctora?|od\.?|cd\.?)\s+", "", s, flags=re.I)
    s = re.sub(r"\s+", " ", s).strip()
    return s

def clean_email(v):
    if v is None or isinstance(v, (datetime, date, float, int)):
        return None
    m = re.search(r"[a-z0-9._%+\-]+@[a-z0-9.\-]+\.[a-z]{2,}", str(v).strip().lower())
    return m.group(0) if m else None

def clean_phone(v):
    if v is None or isinstance(v, (datetime, date)):
        return None
    if isinstance(v, float):
        if v != v or v in (float("inf"), float("-inf")):
            return None
        s = str(int(v))
    elif isinstance(v, int):
        s = str(v)
    else:
        s = re.split(r"[/,;]| y ", str(v))[0]
    digits = re.sub(r"\D", "", s)
    if not digits:
        return None
    if len(digits) == 13 and digits.startswith("521"):
        digits = digits[3:]
    elif len(digits) == 12 and digits.startswith("52"):
        digits = digits[2:]
    elif len(digits) == 11 and digits.startswith("1"):
        digits = digits[1:]
    if len(digits) != 10:
        return None
    return "+52" + digits

def cell_str(v):
    if v is None:
        return ""
    if isinstance(v, float) and v == int(v):
        return str(int(v))
    return str(v).strip()

def load_rows(path, tab):
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    if tab not in wb.sheetnames:
        wb.close()
        raise KeyError(tab)
    rows = []
    for row in wb[tab].iter_rows(values_only=True):
        if any(v is not None and str(v).strip() for v in row):
            rows.append(row)
    wb.close()
    return rows

HEADERY = {"nombre", "pago", "factura", "telefono", "teléfono", "datos", "estado",
           "apellido", "apellidos", "celular", "e-mail", "correo", "direccion",
           "dirección", "doctor", "ortodoncista", "comentarios", "1er mensaje",
           "email", "documentación", "documentacion", "clientes aling", "empresa/consultorio",
           "categoria", "cuidad", "ciudad", "instagram", "curso", "origen del contacto"}

MX_STATES = {
    "aguascalientes", "baja california", "baja california sur", "campeche", "chiapas",
    "chihuahua", "cdmx", "ciudad de mexico", "coahuila", "colima", "durango",
    "estado de mexico", "guanajuato", "guerrero", "hidalgo", "jalisco", "michoacan",
    "morelos", "nayarit", "nuevo leon", "oaxaca", "puebla", "queretaro", "quintana roo",
    "san luis potosi", "sinaloa", "sonora", "tabasco", "tamaulipas", "tlaxcala",
    "veracruz", "yucatan", "zacatecas",
}

def norm_geo(s):
    return re.sub(r"\s+", " ", strip_accents(str(s)).lower()).strip()

def looks_like_name(v, cells_idx=0):
    s = cell_str(v)
    if not s or isinstance(v, (datetime, date, int, float)):
        return None
    if s.lower().strip() in HEADERY or re.search(r"\d|@|http", s):
        return None
    if norm_geo(s) in MX_STATES:  # separadores de sección tipo "TIJUANA"
        return None
    if len(name_tokens(s)) < 2:
        return None
    return clean_name(s)

people = []

def emit(nombre, phone=None, email=None, ciudad=None, estado=None, evento=None,
         fecha=None, fuente=None, competidor=None, casos_anio=None, nota=None,
         warm=False, fecha_contacto=None, interesado=False):
    people.append({
        "nombre": nombre,
        "phone": phone,
        "email": email,
        "ciudad": ciudad or None,
        "estado": estado or None,
        "evento": evento,
        "fecha_evento": fecha,
        "fuente": fuente,   # congreso | curso | evento | b-learning | competidor | outbound
        "competidor": competidor,
        "casos_anio_competidor": casos_anio,
        "nota": nota,
        "warm": warm,                     # true = hubo interacción real (curso/meeting/pago)
        "fecha_contacto": fecha_contacto, # primera fecha del log de outreach
        "interesado": interesado,
    })

# ---------------------------------------------------------------- 1) tabs de eventos (madre)
CURSO_TABS = {
    "AMO 2025": ("2025-02-01", "congreso"),
    "DATOS OBTENIDOS AMO": ("2022-03-04", "congreso"),
    "AMO Monterrey": ("2023-03-01", "congreso"),
    "B Learning 2025": ("2025-04-01", "b-learning"),
    "CURSO MARZO 2025": ("2025-03-20", "curso"),
    "MEETING 4 MARZO": ("2025-03-04", "evento"),
    "DRS CURSO DICIEMBRE 2024": ("2024-12-05", "curso"),
    "DRS CURSO 19  NOVIEMBRE 2024": ("2024-11-19", "curso"),
    "MEETING 20 NOVIEMRE ANDREA": ("2024-11-20", "evento"),
    "MEETING 15 OCTUBRE 2024": ("2024-10-15", "evento"),
    "DRS CURSO 5 de SEPTIEMBRE 2024": ("2024-09-05", "curso"),
    "MEETING TIJUANA 13 de Julio de ": ("2024-07-13", "evento"),
    "MEETING 11 de Julio de 2024": ("2024-07-11", "evento"),
    "CURSO Guadalajara 9 de Julio de": ("2024-07-09", "curso"),
    "DRS CURSO 3 JULIO 2024": ("2024-07-03", "curso"),
    "Masterclass CDMX 16 ABR 24": ("2024-04-16", "evento"),
    "DRS CURSO 10 ABRIL 2024": ("2024-04-10", "curso"),
    "DRS CURSO 13 DICIEMBRE 2023": ("2023-12-13", "curso"),
    "Meeting CDMX 24 NOV 23": ("2023-11-24", "evento"),
    "PRESENCIAL 24 NOV 23": ("2023-11-24", "evento"),
    "DRS CURSO 23 AGOSTO 2023": ("2023-08-23", "curso"),
    "DRS CURSO 22 MARZO 2023": ("2023-03-22", "curso"),
    "DRS CURSO 05 JUNIO 2023": ("2023-06-05", "curso"),
    "DESAYUNO CDMX 27 O2": ("2022-10-27", "evento"),
    "DRS CURSO 25 Nov": ("2022-11-25", "curso"),
    "DRS CURSO 31 ago": ("2022-08-31", "curso"),
    "DRS CURSO 15 jun": ("2022-06-15", "curso"),
    "DRS CURSO 6 de abril": ("2022-04-06", "curso"),
    "Curso universidad": (None, "universidad"),
    "Gira Fabi": (None, "evento"),
}

def src_event_tabs():
    n0 = len(people)
    for tab, (tab_date, tipo) in CURSO_TABS.items():
        try:
            rows = load_rows(F_MADRE, tab)
        except KeyError:
            continue
        # tabs "APELLIDO / NOMBRE" en columnas separadas (cursos con documentación)
        header = [cell_str(v).lower() for v in rows[0]] if rows else []
        split_cols = None
        for i, h in enumerate(header):
            if h in ("apellido", "apellidos"):
                for j, h2 in enumerate(header):
                    if h2 == "nombre" and j != i:
                        split_cols = (i, j)
                        break
        for rix, row in enumerate(rows):
            cells = list(row)
            if split_cols:
                if rix == 0:
                    continue
                ape, nom = cell_str(cells[split_cols[0]]), cell_str(cells[split_cols[1]])
                full = ("%s %s" % (nom, ape)).strip()
                nombre = looks_like_name(full) if full else None
            else:
                nombre = None
                for v in cells[:3]:
                    nombre = looks_like_name(v)
                    if nombre:
                        break
            if not nombre:
                continue
            email = next((clean_email(v) for v in cells if clean_email(v)), None)
            phone = None
            for v in cells[1:]:
                s = cell_str(v)
                if re.search(r"\$|%", s):
                    continue
                p = clean_phone(v)
                if p:
                    phone = p
                    break
            estado = next((cell_str(v) for v in cells if norm_geo(v) in MX_STATES), None)
            # Gira Fabi: pares de columnas (email, nombre, tel) x2 — segunda persona
            emit(nombre, phone, email, None, estado, tab.strip(), tab_date, tipo, warm=True)
            if tab == "Gira Fabi" and len(cells) > 7:
                nom2 = looks_like_name(cells[6])
                if nom2:
                    emit(nom2, clean_phone(cells[7]) if len(cells) > 7 else None,
                         clean_email(cells[5]), None, None, tab.strip(), tab_date, tipo,
                         warm=True)
    print("eventos: %d personas" % (len(people) - n0))

# ---------------------------------------------------------------- 2) competidores (madre)
def src_competitor_tab(path, tab, marca, name_cols=(0, 1), city_col=3, phone_cols=(6, 7)):
    n0 = len(people)
    try:
        rows = load_rows(path, tab)
    except KeyError:
        print("  (tab %s no encontrado)" % tab)
        return
    for row in rows:
        cells = list(row)
        nom = cell_str(cells[name_cols[0]]) if len(cells) > name_cols[0] else ""
        ape = cell_str(cells[name_cols[1]]) if len(cells) > name_cols[1] else ""
        full = ("%s %s" % (nom, ape)).strip()
        nombre = looks_like_name(full) if full else None
        if not nombre:
            continue
        phone = None
        for c in phone_cols:
            if len(cells) > c:
                phone = clean_phone(cells[c])
                if phone:
                    break
        email = next((clean_email(v) for v in cells if clean_email(v)), None)
        ciudad = cell_str(cells[city_col]) if len(cells) > city_col else None
        emit(nombre, phone, email, ciudad, None, None, None, "competidor",
             competidor=marca)
    print("%s: %d personas" % (marca, len(people) - n0))

# ---------------------------------------------------------------- 3) Invisalign por estado
ALIGN_SKIP_TABS = {"Casos por categoria", "CLINCAS GRANDES"}

def src_invisalign():
    n0 = len(people)
    wb = openpyxl.load_workbook(F_INVIS, read_only=True, data_only=True)
    tabs = [t for t in wb.sheetnames if t not in ALIGN_SKIP_TABS]
    wb.close()
    for tab in tabs:
        rows = load_rows(F_INVIS, tab)
        for rix, row in enumerate(rows):
            if rix == 0:
                continue
            cells = list(row)
            nombre = looks_like_name(cells[0]) if cells else None
            if not nombre:
                continue
            phone = clean_phone(cells[2]) if len(cells) > 2 else None
            # máximo "Cantidad de Casos Q*" numérico = techo anual demostrado en Align
            casos = None
            for v in cells[3:]:
                if isinstance(v, (int, float)) and 0 < float(v) < 500:
                    casos = max(casos or 0, float(v))
            emit(nombre, phone, None, tab.strip().title(), None, None, None,
                 "competidor", competidor="Invisalign", casos_anio=casos)
    print("Invisalign: %d personas" % (len(people) - n0))

# ---------------------------------------------------------------- 4) Acredi Norte extras
def src_norte():
    n0 = len(people)
    # Potenciales Align: col1 nombre, col2/3 tel, col4 email, col7 ciudad, col8 estado
    try:
        for row in load_rows(F_NORTE, "Potenciales Align"):
            cells = list(row)
            nombre = looks_like_name(cells[1]) if len(cells) > 1 else None
            if not nombre:
                continue
            phone = (clean_phone(cells[2]) if len(cells) > 2 else None) or \
                    (clean_phone(cells[3]) if len(cells) > 3 else None)
            email = next((clean_email(v) for v in cells if clean_email(v)), None)
            ciudad = cell_str(cells[7]) if len(cells) > 7 else None
            estado = cell_str(cells[8]) if len(cells) > 8 else None
            emit(nombre, phone, email, ciudad, estado, None, None, "competidor",
                 competidor="Invisalign", nota="Potenciales Align (Norte)")
    except KeyError:
        pass
    # Cryst aligner (Norte) — mismo formato que el de la madre
    src_competitor_tab(F_NORTE, "Cryst aligner ", "Cryst Aligner")
    # 'A' — lista miscelánea formato Aliwell
    src_competitor_tab(F_NORTE, "A", "(sin marca)", name_cols=(0, 1), city_col=3,
                       phone_cols=(6,))
    print("norte extras: total acumulado %d" % (len(people) - n0))

# ---------------------------------------------------------------- 5) B-Learning Bono
def src_blearning():
    n0 = len(people)
    try:
        for rix, row in enumerate(load_rows(F_BLEARN, "Inscriptos")):
            if rix < 2:
                continue
            cells = list(row)
            nombre = looks_like_name(cells[0]) if cells else None
            if not nombre:
                continue
            acreditado = cell_str(cells[9]).upper() if len(cells) > 9 else ""
            emit(nombre, clean_phone(cells[1]) if len(cells) > 1 else None,
                 next((clean_email(v) for v in cells if clean_email(v)), None),
                 None, None, "B-Learning Bono", "2025-04-16", "b-learning",
                 nota="acreditado=%s" % (acreditado or "?"), warm=True)
    except KeyError:
        pass
    try:
        for rix, row in enumerate(load_rows(F_BLEARN, "Contactados")):
            if rix == 0:
                continue
            cells = list(row)
            nombre = looks_like_name(cells[0]) if cells else None
            if not nombre:
                continue
            categoria = cell_str(cells[1]) if len(cells) > 1 else ""
            if categoria.lower().startswith("acreditado"):
                continue  # ya es cliente; el enrichment ya los cruzó
            emit(nombre, clean_phone(cells[5]) if len(cells) > 5 else None,
                 None, None, cell_str(cells[6]) if len(cells) > 6 else None,
                 "B-Learning contactados", None, "b-learning",
                 nota=cell_str(cells[9])[:120] if len(cells) > 9 else None, warm=True)
    except KeyError:
        pass
    print("b-learning: %d personas" % (len(people) - n0))

# ---------------------------------------------------------------- 6) bases de outreach (madre)
def src_base_contactados():
    """BASE DRS CONTACTADOS: base de outbound con procedencia y 1 ronda de contacto."""
    n0 = len(people)
    try:
        rows = load_rows(F_MADRE, "BASE DRS CONTACTADOS")
    except KeyError:
        return
    for rix, row in enumerate(rows):
        if rix == 0:
            continue
        cells = list(row)
        def at(i):
            return cell_str(cells[i]) if len(cells) > i else ""
        full = ("%s %s" % (at(0), at(1))).strip()
        nombre = looks_like_name(full) if full else None
        if not nombre:
            continue
        phone = (clean_phone(cells[6]) if len(cells) > 6 else None) or \
                (clean_phone(cells[7]) if len(cells) > 7 else None)
        email = clean_email(cells[8]) if len(cells) > 8 else None
        if not phone and not email:
            continue  # sin canal no es accionable
        origen = at(10).upper()
        marca = None
        for m, tag in (("ALIWELL", "Aliwell"), ("CRYST", "Cryst Aligner"),
                       ("INVIS", "Invisalign"), ("CLEAR", "ClearCorrect"),
                       ("SMART", "SmartLigner")):
            if m in origen:
                marca = tag
                break
        estado_c = at(17)
        quien = at(18)
        f1 = None
        if len(cells) > 19 and isinstance(cells[19], (datetime, date)):
            f1 = cells[19].date().isoformat() if isinstance(cells[19], datetime) else cells[19].isoformat()
        contactado = bool(estado_c or quien or f1)
        nota = " · ".join(x for x in [
            ("origen: %s" % at(10)) if at(10) else None,
            estado_c or None,
            ("contactó %s" % quien) if quien else None,
        ] if x) or None
        emit(nombre, phone, email, at(3), None, None, None,
             "outbound", competidor=marca, nota=nota, warm=contactado,
             fecha_contacto=f1)
    print("BASE DRS CONTACTADOS: %d personas" % (len(people) - n0))

def src_contactados_2022():
    """DRES CONTACTADOS SEP 2022: log de gestión con hasta 3+ rondas comentadas."""
    n0 = len(people)
    try:
        rows = load_rows(F_MADRE, "DRES CONTACTADOS SEP 2022")
    except KeyError:
        return
    header = [cell_str(v).lower() for v in rows[0]]
    # posiciones de las rondas: cada 'quien lo contacto' abre una ronda
    rounds = []
    for i, h in enumerate(header):
        if h.startswith("quien lo contacto"):
            rounds.append(i)
    for rix, row in enumerate(rows):
        if rix == 0:
            continue
        cells = list(row)
        def at(i):
            return cell_str(cells[i]) if len(cells) > i else ""
        full = ("%s %s" % (at(0), at(1))).strip()
        nombre = looks_like_name(full) if full else None
        if not nombre:
            continue
        phone = (clean_phone(cells[6]) if len(cells) > 6 else None) or \
                (clean_phone(cells[7]) if len(cells) > 7 else None)
        email = clean_email(cells[8]) if len(cells) > 8 else None
        if not phone and not email:
            continue
        origen = at(9)
        marca = None
        for m, tag in (("ALIWELL", "Aliwell"), ("CRYST", "Cryst Aligner"),
                       ("INVIS", "Invisalign")):
            if m in origen.upper():
                marca = tag
                break
        notas, f1, interesado, contactado = [], None, False, False
        for start in rounds[:3]:
            quien = at(start)
            fecha_v = cells[start + 1] if len(cells) > start + 1 else None
            fecha = None
            if isinstance(fecha_v, (datetime, date)):
                fecha = (fecha_v.date() if isinstance(fecha_v, datetime) else fecha_v).isoformat()
            quepaso = at(start + 2)
            comentario = at(start + 4)
            if not (quien or fecha or quepaso or comentario):
                continue
            contactado = True
            f1 = f1 or fecha
            if "interesado" in quepaso.lower():
                interesado = True
            notas.append(" ".join(x for x in [fecha or "", quien, quepaso,
                                              ("— " + comentario) if comentario else ""] if x))
        estado_c = at(10)
        nota_full = " · ".join(x for x in ([("origen: %s" % origen) if origen else None,
                                            estado_c or None] + notas) if x)
        emit(nombre, phone, email, at(3) if at(3).upper() != "PROXIMO CURSO" else None,
             None, None, None, "outbound", competidor=marca,
             nota=nota_full[:400] or None, warm=contactado, fecha_contacto=f1,
             interesado=interesado)
    print("DRES CONTACTADOS SEP 2022: %d personas" % (len(people) - n0))

def src_sheet30():
    n0 = len(people)
    try:
        rows = load_rows(F_MADRE, "Sheet30")
    except KeyError:
        return
    for rix, row in enumerate(rows):
        if rix == 0:
            continue
        cells = list(row)
        nombre = looks_like_name(cells[2]) if len(cells) > 2 else None
        if not nombre:
            continue
        emit(nombre, clean_phone(cells[1]) if len(cells) > 1 else None,
             None, None, None, None, None, "competidor", competidor="SmartLigner")
    print("Sheet30/SmartLigner: %d personas" % (len(people) - n0))

def src_mail_clearcorrect():
    """MAIL DOCTORES = mails de ortodoncistas ClearCorrect (sin nombre) →
    archivo aparte para etiquetar por email, no crea fichas."""
    emails = []
    try:
        for row in load_rows(F_MADRE, "MAIL DOCTORES"):
            for v in row:
                e = clean_email(v)
                if e:
                    emails.append(e)
    except KeyError:
        pass
    emails = sorted(set(emails))
    with open(os.path.join(DATA, "clearcorrect_emails.json"), "w") as f:
        json.dump(emails, f, indent=0)
    print("ClearCorrect emails: %d" % len(emails))

# ---------------------------------------------------------------- main
def main():
    src_event_tabs()
    src_competitor_tab(F_MADRE, "Aliwell", "Aliwell", name_cols=(0, 1), city_col=3,
                       phone_cols=(6,))
    src_competitor_tab(F_MADRE, "Ortodoncistas Cryst", "Cryst Aligner")
    src_invisalign()
    src_norte()
    src_blearning()
    src_base_contactados()
    src_contactados_2022()
    src_sheet30()
    src_mail_clearcorrect()

    # ---- dedupe interno: tel > email > nombre normalizado (prioriza warm y con más datos)
    def score(p):
        return (1 if p["warm"] else 0, 1 if p["phone"] else 0, 1 if p["email"] else 0,
                1 if p["fecha_evento"] else 0)
    people.sort(key=score, reverse=True)
    seen_phone, seen_email, seen_name = {}, {}, {}
    out = []
    for p in people:
        keyp = p["phone"]
        keye = p["email"]
        keyn = " ".join(sorted(name_tokens(p["nombre"])))
        winner = (keyp and seen_phone.get(keyp)) or (keye and seen_email.get(keye)) \
                 or (keyn and seen_name.get(keyn))
        if winner:
            # fusionar en el ganador lo que a este le falte
            for k in ("phone", "email", "ciudad", "estado", "competidor",
                      "casos_anio_competidor", "fecha_contacto", "evento",
                      "fecha_evento"):
                if not winner.get(k) and p.get(k):
                    winner[k] = p[k]
            if p.get("nota"):
                winner["nota"] = ((winner.get("nota") or "") + " | " + p["nota"])[:400]
            winner["warm"] = winner["warm"] or p["warm"]
            winner["interesado"] = winner.get("interesado") or p.get("interesado")
            continue
        if keyp:
            seen_phone[keyp] = p
        if keye:
            seen_email[keye] = p
        if keyn:
            seen_name[keyn] = p
        out.append(p)

    with open(os.path.join(DATA, "prospectos_fuentes.json"), "w") as f:
        json.dump(out, f, ensure_ascii=False, indent=1)

    by_fuente = {}
    con_tel = sum(1 for p in out if p["phone"])
    for p in out:
        by_fuente[p["fuente"]] = by_fuente.get(p["fuente"], 0) + 1
    print("\nTOTAL únicos: %d (de %d brutos) · con teléfono: %d (%d%%)"
          % (len(out), len(people), con_tel, 100 * con_tel // max(1, len(out))))
    for k, v in sorted(by_fuente.items(), key=lambda x: -x[1]):
        print("  %-12s %d" % (k, v))

if __name__ == "__main__":
    main()
