#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
parse_enrichment.py — CRM MX import enrichment parser.

Reads local XLSX copies of the Google Sheets corpus + the Noloco doctor base and emits:
  crm-mx/data/enrichment.json        (per matched Noloco doctor)
  crm-mx/data/payments.json          (one record per dated payment in the admin ledger)
  crm-mx/data/enrichment_report.md   (coverage + anomalies)

Python 3.9 / openpyxl. Read-only on the sheets.
"""
import json
import os
import re
import sys
import unicodedata
from datetime import datetime, date

import openpyxl

# ---------------------------------------------------------------- paths
SHEETS_DIR = ("/private/tmp/claude-501/-Users-franciscobasilico-Desktop-Claude-Periskope/"
              "dfb9ce39-9fac-4ca6-a9b4-daf01a381467/scratchpad/crm_sheets")
NOLOCO_JSON = "/Users/franciscobasilico/dev/Periskope/gestion-mx/data/noloco_mx.json"
OUT_DIR = "/Users/franciscobasilico/dev/Periskope/crm-mx/data"

F_ADMIN = "12n4w566gJmHa1ky73dmsRwRoHMorYemq95BIOCSKXog.xlsx"   # Administración México
F_BLEARN = "19epMNc-KzHt3hobw8cEjcvSdTLe8WQhwFlPgL40P3Dw.xlsx"  # B-Learning Bono
F_NORTE = "1Zv_G9OyPDvyv-Aft4-eYLTvFCJ1acBobp90uFXNso5g.xlsx"   # Acreditados Zona Norte
F_MADRE = "1db8uSRkPo68hCx1fpD-8a9qEMrvbxIZGzovxnMudZA0.xlsx"   # Mexico Pancho_Juan (madre)
F_CRM = "1dejMjxo-nhBYXHsphe22qQxfv2pZbaYeek2UsQqpC7E.xlsx"     # CRM_KS_MX
F_ENTREGAS = "1lQEFgW0JRwFosHpBKKjkF-OjS9J9Q3ThfhP01WiU3AE.xlsx"  # Entregas MEXICO

# ---------------------------------------------------------------- normalization helpers
TITLE_TOKENS = {"dr", "dra", "drs", "dres", "doctor", "doctora", "cks", "imed", "od", "cd", "mtro", "mtra", "esp"}
PARTICLES = {"de", "del", "la", "las", "los", "y", "e", "da", "do", "van", "der"}

def strip_accents(s):
    return "".join(c for c in unicodedata.normalize("NFD", s) if unicodedata.category(c) != "Mn")

def name_tokens(raw):
    """Normalized significant token set for a doctor name."""
    if raw is None:
        return frozenset()
    s = strip_accents(str(raw)).lower()
    s = re.sub(r"\(.*?\)", " ", s)           # (CKS) (IMED) ...
    s = re.sub(r"[^a-z\s]", " ", s)
    toks = [t for t in s.split() if t not in TITLE_TOKENS and t not in PARTICLES and len(t) > 1]
    return frozenset(toks)

def clean_email(v):
    if v is None or isinstance(v, (datetime, date, float, int)):
        return None
    s = str(v).strip().lower()
    m = re.search(r"[a-z0-9._%+\-]+@[a-z0-9.\-]+\.[a-z]{2,}", s)
    return m.group(0) if m else None

PHONE_SKIPPED_DATETIMES = [0]

def clean_phone(v):
    """Normalize to +52XXXXXXXXXX (MX national 10 digits). None if unusable."""
    if v is None:
        return None
    if isinstance(v, (datetime, date)):
        PHONE_SKIPPED_DATETIMES[0] += 1
        return None
    if isinstance(v, float):
        if v != v or v in (float("inf"), float("-inf")):
            return None
        s = str(int(v))
    elif isinstance(v, int):
        s = str(v)
    else:
        s = str(v)
        # multiple numbers: take the first chunk
        s = re.split(r"[/,;]| y ", s)[0]
    digits = re.sub(r"\D", "", s)
    if not digits:
        return None
    if len(digits) == 13 and digits.startswith("521"):
        digits = digits[3:]
    elif len(digits) == 12 and digits.startswith("52"):
        digits = digits[2:]
    elif len(digits) == 11 and digits.startswith("1"):
        digits = digits[1:]
    if len(digits) != 10:
        return None
    return "+52" + digits

def to_date(v):
    if v is None:
        return None
    if isinstance(v, datetime):
        return v.date().isoformat()
    if isinstance(v, date):
        return v.isoformat()
    s = str(v).strip()
    for fmt in ("%Y-%m-%d %H:%M:%S", "%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y"):
        try:
            return datetime.strptime(s[:19], fmt).date().isoformat()
        except ValueError:
            continue
    return None

def to_amount(v):
    if v is None:
        return None
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).strip().replace("$", "").replace(",", "").replace(" ", "")
    if not s or s in ("-", "—"):
        return None
    try:
        return float(s)
    except ValueError:
        return None

def cell_str(v):
    if v is None:
        return ""
    if isinstance(v, float) and v == int(v):
        return str(int(v))
    return str(v).strip()

# ---------------------------------------------------------------- geography
MX_STATES = {
    "aguascalientes", "baja california", "baja california sur", "campeche", "chiapas", "chihuahua",
    "coahuila", "colima", "durango", "guanajuato", "guerrero", "hidalgo", "jalisco", "michoacan",
    "morelos", "nayarit", "nuevo leon", "oaxaca", "puebla", "queretaro", "quintana roo",
    "san luis potosi", "sinaloa", "sonora", "tabasco", "tamaulipas", "tlaxcala", "veracruz",
    "yucatan", "zacatecas", "cdmx", "ciudad de mexico", "estado de mexico", "edo mex", "edomex",
    "estado mexico", "mexico",
}
CITY_TO_STATE = {
    "mexicali": "Baja California", "tijuana": "Baja California", "ensenada": "Baja California",
    "monterrey": "Nuevo León", "saltillo": "Coahuila", "torreon": "Coahuila",
    "cd juarez": "Chihuahua", "ciudad juarez": "Chihuahua", "juarez": "Chihuahua",
    "chihuahua": "Chihuahua", "hermosillo": "Sonora", "culiacan": "Sinaloa", "mazatlan": "Sinaloa",
    "guadalajara": "Jalisco", "zapopan": "Jalisco", "toluca": "Estado de México",
    "metepec": "Estado de México", "cuernavaca": "Morelos", "merida": "Yucatán",
    "cancun": "Quintana Roo", "puebla": "Puebla", "queretaro": "Querétaro", "leon": "Guanajuato",
    "morelia": "Michoacán", "oaxaca": "Oaxaca", "villahermosa": "Tabasco", "tampico": "Tamaulipas",
    "veracruz": "Veracruz", "xalapa": "Veracruz", "pachuca": "Hidalgo", "durango": "Durango",
    "la paz": "Baja California Sur", "acapulco": "Guerrero", "tuxtla gutierrez": "Chiapas",
    "san luis potosi": "San Luis Potosí", "aguascalientes": "Aguascalientes",
    "campeche": "Campeche", "tlaxcala": "Tlaxcala", "colima": "Colima", "tepic": "Nayarit",
    "zacatecas": "Zacatecas", "cdmx": "CDMX", "coyoacan": "CDMX", "texcoco": "Estado de México",
}
ZONA_CDMX = {"cdmx", "ciudad de mexico", "estado de mexico", "edo mex", "edomex", "estado mexico",
             "mexico", "toluca", "metepec", "texcoco", "coyoacan"}
ZONA_NORTE = {"baja california", "baja california sur", "chihuahua", "cd juarez", "ciudad juarez",
              "juarez", "nuevo leon", "monterrey", "sonora", "sinaloa", "coahuila", "tamaulipas",
              "mexicali", "tijuana", "ensenada", "durango", "saltillo", "torreon", "hermosillo",
              "culiacan", "mazatlan", "tampico", "la paz"}
ZONA_SUR = {"puebla", "oaxaca", "guerrero", "chiapas", "veracruz", "yucatan", "merida",
            "quintana roo", "tabasco", "campeche", "cancun", "acapulco", "villahermosa",
            "xalapa", "tuxtla gutierrez"}

def norm_geo(s):
    return re.sub(r"\s+", " ", strip_accents(str(s)).lower().replace(".", " ").replace(",", " ")).strip()

def split_city_state(raw):
    """A 'Region'/'Localidad' value can be a city or a state. Return (city, state)."""
    v = cell_str(raw)
    if not v or re.fullmatch(r"[\d.]+", v):
        return None, None
    n = norm_geo(v)
    if n in MX_STATES:
        return None, v.strip()
    st = CITY_TO_STATE.get(n)
    return v.strip(), st

def derive_zona(city, state):
    for cand in (state, city):
        if not cand:
            continue
        n = norm_geo(cand)
        if n in ZONA_CDMX:
            return "CDMX"
        if n in ZONA_NORTE:
            return "Norte"
        if n in ZONA_SUR:
            return "Sur"
    if city or state:
        return "Foráneos"
    return None

# ---------------------------------------------------------------- competitor brands
def normalize_brand(raw):
    if not raw:
        return None
    n = strip_accents(str(raw)).lower()
    if "invisal" in n or "invisaling" in n or "align" == n.strip():
        return "Invisalign"
    if "b360" in n or "cryst" in n:
        return "Cryst"
    if "aliwell" in n:
        return "Aliwell"
    if "ormco" in n or "spark" in n:
        return "Spark"
    if "smartee" in n:
        return "Smartee"
    if "moons" in n:
        return "moons"
    return None

# ---------------------------------------------------------------- doctor base + matcher
class DoctorBase:
    def __init__(self, path):
        data = json.load(open(path))
        self.doctors = {}
        self.case_counts = {}
        for c in data["casos"]:
            d = c.get("doctores") or {}
            if not d.get("id"):
                continue
            self.case_counts[d["id"]] = self.case_counts.get(d["id"], 0) + 1
            if d["id"] not in self.doctors:
                self.doctors[d["id"]] = {
                    "id": d["id"],
                    "nombre": (d.get("nombre") or "").strip(),
                    "email": (d.get("email") or "").strip().lower() or None,
                }
        all_tokens = {i: name_tokens(d["nombre"]) for i, d in self.doctors.items()}
        # duplicate doctors in Noloco (identical normalized names, e.g. accent variants):
        # match against the canonical one (most casos) only.
        groups = {}
        for i, t in all_tokens.items():
            groups.setdefault(t, []).append(i)
        self.duplicates = {}  # dup_id -> canonical_id
        for t, ids in groups.items():
            if len(ids) > 1 and t:
                ids.sort(key=lambda i: -self.case_counts.get(i, 0))
                for dup in ids[1:]:
                    self.duplicates[dup] = ids[0]
        self.tokens = {i: t for i, t in all_tokens.items() if i not in self.duplicates}
        self.by_email = {d["email"]: i for i, d in self.doctors.items() if d["email"]}
        for dup, canon in self.duplicates.items():
            em = self.doctors[dup]["email"]
            if em and self.by_email.get(em) == dup:
                self.by_email[em] = canon
        self._cache = {}
        self.ambiguous = {}  # raw_key -> [candidate names]

    def match_email(self, email):
        if not email:
            return None
        return self.by_email.get(email)

    def match_name(self, raw):
        """Token-set fuzzy match. Returns doctor id or None. Ambiguous -> None (recorded)."""
        key = " ".join(sorted(name_tokens(raw)))
        if not key:
            return None
        if key in self._cache:
            return self._cache[key]
        toks = frozenset(key.split())
        scored = []
        for did, dtoks in self.tokens.items():
            if not dtoks:
                continue
            inter = len(toks & dtoks)
            if inter < 2:
                continue
            jac = inter / len(toks | dtoks)
            # subset bonus: all tokens of the shorter side contained in the longer side
            subset = toks <= dtoks or dtoks <= toks
            if jac >= 0.5 or (subset and inter >= 2):
                scored.append((inter, jac + (0.15 if subset else 0.0), did))
        result = None
        if scored:
            scored.sort(reverse=True)
            if len(scored) > 1:
                (i1, s1, _), (i2, s2, _) = scored[0], scored[1]
                # a candidate covering strictly more of the name's tokens wins outright;
                # equal coverage with close scores = ambiguous, leave unmatched
                if i1 == i2 and s1 - s2 < 0.12:
                    self.ambiguous[key] = [self.doctors[s[2]]["nombre"] for s in scored[:3]]
                else:
                    result = scored[0][2]
            else:
                result = scored[0][2]
        self._cache[key] = result
        return result

    def match(self, raw_name, email=None):
        did = self.match_email(clean_email(email)) if email else None
        if did:
            return did, "email"
        did = self.match_name(raw_name)
        return (did, "name") if did else (None, None)

# ---------------------------------------------------------------- enrichment accumulator
FIELDS = ("phone", "whatsapp", "email", "city", "state", "accredited_at", "comercial_asignado")

class Enrichment:
    def __init__(self, base):
        self.base = base
        self.rec = {i: {f: None for f in FIELDS} for i in base.doctors}
        self.src = {i: {} for i in base.doctors}
        self.brands = {i: set() for i in base.doctors}

    def set(self, did, field, value, source):
        if value in (None, ""):
            return
        if self.rec[did][field] is None:
            self.rec[did][field] = value
            self.src[did][field] = source

    def add_brand(self, did, brand, source):
        if brand and brand not in self.brands[did]:
            self.brands[did].add(brand)
            self.src[did].setdefault("competitor_brands", source)

# ---------------------------------------------------------------- sheet loading
_wb_cache = {}
_rows_cache = {}

def load_rows(fname, tab):
    key = (fname, tab)
    if key not in _rows_cache:
        if fname not in _wb_cache:
            _wb_cache[fname] = openpyxl.load_workbook(
                os.path.join(SHEETS_DIR, fname), read_only=True, data_only=True)
        wb = _wb_cache[fname]
        if tab not in wb.sheetnames:
            raise KeyError(tab)
        _rows_cache[key] = [list(r) for r in wb[tab].iter_rows(values_only=True)]
    return _rows_cache[key]

def find_header(rows, must_contain, max_scan=6):
    """Locate header row index by required column names (case/space-insensitive)."""
    want = [w.lower() for w in must_contain]
    for i, row in enumerate(rows[:max_scan]):
        cells = [cell_str(c).lower().strip() for c in row]
        if all(any(w == c or w in c for c in cells if c) for w in want):
            return i
    return None

def col_index(header_row, name, occurrence=0):
    name = name.lower().strip()
    hits = [i for i, c in enumerate(header_row) if cell_str(c).lower().strip() == name]
    if len(hits) > occurrence:
        return hits[occurrence]
    # fallback: contains
    hits = [i for i, c in enumerate(header_row) if name in cell_str(c).lower()]
    return hits[occurrence] if len(hits) > occurrence else None

# ================================================================ SOURCES

def src_acredi_norte(enr, base, stats):
    """Priority 1 phones/emails + city/state (prio 3) + accredited_at (prio 1)."""
    rows = load_rows(F_NORTE, "Acredi Norte")
    n = 0
    for row in rows:
        if len(row) < 15:
            row = list(row) + [None] * (15 - len(row))
        name = cell_str(row[4])
        if not name or len(name_tokens(name)) < 2:
            continue
        phone = clean_phone(row[5])
        email = clean_email(row[6])
        city_raw = cell_str(row[9])
        state_raw = cell_str(row[10])
        acc = to_date(row[14])
        did, how = base.match(name, email)
        if not did:
            continue
        n += 1
        enr.set(did, "phone", phone, "Acredi Norte")
        enr.set(did, "email", email, "Acredi Norte")
        city = None if re.fullmatch(r"[\d.]*", city_raw) else city_raw.strip()
        state = None if re.fullmatch(r"[\d.]*", state_raw) else state_raw.strip()
        if city and not state:
            state = CITY_TO_STATE.get(norm_geo(city))
        enr.set(did, "_geo3_city", None, "")  # placeholder no-op
        # stash as priority-3 geo (applied later)
        stats["geo3"].setdefault(did, (city, state, "Acredi Norte"))
        enr.set(did, "accredited_at", acc, "Acredi Norte")
    stats["matched_acredi"] = n

def src_blearning(enr, base, stats):
    """Priority 2 phones: B-Learning Bono 'Contactados' + 'Inscriptos'."""
    rows = load_rows(F_BLEARN, "Contactados")
    hi = find_header(rows, ["Ortodoncista", "celular"])
    if hi is None:
        return
    h = rows[hi]
    c_name, c_cel = col_index(h, "ortodoncista"), col_index(h, "celular")
    n = 0
    for row in rows[hi + 1:]:
        name = cell_str(row[c_name]) if c_name < len(row) else ""
        if not name:
            continue
        did, _ = base.match(name)
        if not did:
            continue
        n += 1
        enr.set(did, "phone", clean_phone(row[c_cel] if c_cel < len(row) else None), "B-Learning Contactados")
    stats["matched_contactados"] = n

    rows = load_rows(F_BLEARN, "Inscriptos")
    hi = find_header(rows, ["DOCTOR", "TELEFONO"])
    if hi is None:
        return
    h = rows[hi]
    c_name, c_tel = col_index(h, "doctor"), col_index(h, "telefono")
    for row in rows[hi + 2:]:  # skip CORREO/DIRECCIÓN subheader row
        name = cell_str(row[c_name]) if c_name < len(row) else ""
        if not name:
            continue
        did, _ = base.match(name)
        if not did:
            continue
        enr.set(did, "phone", clean_phone(row[c_tel] if c_tel < len(row) else None), "B-Learning Inscriptos")
        for v in row:
            em = clean_email(v)
            if em:
                enr.set(did, "email", em, "B-Learning Inscriptos")
                break

# Curso/AMO/meeting tabs of the madre sheet, with inferred event dates (None = unknown).
CURSO_TABS = {
    "AMO 2025": "2025-02-01",
    "B Learning 2025": "2025-04-01",
    "CURSO MARZO 2025": "2025-03-20",
    "MEETING 4 MARZO": "2025-03-04",
    "DRS CURSO DICIEMBRE 2024": "2024-12-05",
    "DRS CURSO 19  NOVIEMBRE 2024": "2024-11-19",
    "MEETING 20 NOVIEMRE ANDREA": "2024-11-20",
    "MEETING 15 OCTUBRE 2024": "2024-10-15",
    "DRS CURSO 5 de SEPTIEMBRE 2024": "2024-09-05",
    "MEETING TIJUANA 13 de Julio de ": "2024-07-13",
    "MEETING 11 de Julio de 2024": "2024-07-11",
    "CURSO Guadalajara 9 de Julio de": "2024-07-09",
    "DRS CURSO 3 JULIO 2024": "2024-07-03",
    "Masterclass CDMX 16 ABR 24": "2024-04-16",
    "DRS CURSO 10 ABRIL 2024": "2024-04-10",
    "DRS CURSO 13 DICIEMBRE 2023": "2023-12-13",
    "Meeting CDMX 24 NOV 23": "2023-11-24",
    "PRESENCIAL 24 NOV 23": "2023-11-24",
    "DRS CURSO 23 AGOSTO 2023": "2023-08-23",
    "DRS CURSO 22 MARZO 2023": "2023-03-22",
    "DRS CURSO 05 JUNIO 2023": "2023-06-05",
    "DESAYUNO CDMX 27 O2": "2022-10-27",
    "DRS CURSO 25 Nov": "2022-11-25",
    "DRS CURSO 31 ago": "2022-08-31",
    "DRS CURSO 15 jun": "2022-06-15",
    "DRS CURSO 6 de abril": "2022-04-06",
    "Curso universidad": None,
    "DATOS OBTENIDOS AMO": "2022-03-04",
    "AMO Monterrey": "2023-03-01",
    "Gira Fabi": None,
}

HEADERY = {"nombre", "pago", "factura", "telefono", "datos", "estado", "apellido", "celular",
           "e-mail", "correo", "direccion", "dirección", "doctor", "comentarios", "1er mensaje",
           "email", "centro de escaneo", "documentación", "documentacion", "redes sociales"}

MX_STATE_HINTS = MX_STATES | {"cdmx ", "saltillo", "tijuana", "monterrey", "nuevo leon monterrey"}

def src_curso_tabs(enr, base, stats):
    """Priority 3 phones/emails; ESTADO as prio-4 geo; tab date as prio-3 accredited_at."""
    n = 0
    for tab, tab_date in CURSO_TABS.items():
        try:
            rows = load_rows(F_MADRE, tab)
        except KeyError:
            continue
        for row in rows:
            cells = list(row)
            # name: first non-empty string cell that isn't a header keyword / date / number
            name = None
            for v in cells[:3]:
                s = cell_str(v)
                if (s and not isinstance(v, (datetime, date, int, float))
                        and s.lower().strip() not in HEADERY
                        and len(name_tokens(s)) >= 2 and not re.search(r"\d", s)):
                    name = s
                    break
            if not name:
                continue
            email = next((clean_email(v) for v in cells if clean_email(v)), None)
            did, how = base.match(name, email)
            if not did:
                continue
            n += 1
            phone = None
            for v in cells[1:]:
                if isinstance(v, (datetime, date)):
                    continue
                s = cell_str(v)
                if re.search(r"\$|%", s):
                    continue
                p = clean_phone(v)
                if p:
                    phone = p
                    break
            estado = None
            for v in cells:
                s = cell_str(v)
                if s and norm_geo(s) in MX_STATE_HINTS:
                    estado = s
                    break
            enr.set(did, "phone", phone, "curso:%s" % tab)
            enr.set(did, "email", email, "curso:%s" % tab)
            if estado:
                stats["geo4"].setdefault(did, (None, estado.strip(), "curso:%s" % tab))
            if tab_date:
                stats["acc3"].setdefault(did, (tab_date, "curso:%s" % tab))
    stats["matched_cursos"] = n

def src_big_bases(enr, base, stats):
    """Priority 4 phones/emails + FUENTE competitor brands from BASE DE DATOS COMPLETA."""
    # header taken from 'Ortodoncistas no acreditados Ba'
    C_NOM, C_APE, C_CIUDAD, C_TEL, C_CEL, C_CORREO, C_FUENTE = 0, 1, 3, 6, 7, 8, 10
    n = 0
    for tab, has_header in (("BASE DE DATOS COMPLETA", False),
                            ("Ortodoncistas no acreditados Ba", True),
                            ("Ortodoncistas no acreditados", False),
                            ("BASE DE DATOS NO ACREDITADOS", False),
                            ("BASE DRS CONTACTADOS", False),
                            ("DRES CONTACTADOS SEP 2022", False)):
        try:
            rows = load_rows(F_MADRE, tab)
        except KeyError:
            continue
        for i, row in enumerate(rows):
            if has_header and i == 0:
                continue
            cells = list(row) + [None] * max(0, 12 - len(row))
            full = (cell_str(cells[C_NOM]) + " " + cell_str(cells[C_APE])).strip()
            if len(name_tokens(full)) < 2:
                continue
            email = clean_email(cells[C_CORREO])
            did, how = base.match(full, email)
            if not did:
                continue
            n += 1
            src = "base:%s" % tab
            phone = clean_phone(cells[C_CEL]) or clean_phone(cells[C_TEL])
            enr.set(did, "phone", phone, src)
            enr.set(did, "email", email, src)
            city, state = split_city_state(cells[C_CIUDAD])
            if city or state:
                stats["geo5"].setdefault(did, (city, state, src))
            brand = normalize_brand(cells[C_FUENTE])
            enr.add_brand(did, brand, src)
    stats["matched_bases"] = n

def src_competitor_tabs(enr, base, stats):
    """Brand-specific tabs."""
    specs = [
        (F_MADRE, "Ortodoncistas Cryst", "Cryst", lambda r: (cell_str(r[0]) + " " + cell_str(r[1])).strip(), 1),
        (F_MADRE, "Cryst aligner ", "Cryst", lambda r: (cell_str(r[0]) + " " + cell_str(r[1])).strip(), 1),
        (F_MADRE, "Aliwell", "Aliwell", lambda r: (cell_str(r[0]) + " " + cell_str(r[1])).strip(), 0),
        (F_NORTE, "Drs invisa BC", "Invisalign", lambda r: cell_str(r[1]) if len(r) > 1 else "", 1),
        (F_NORTE, "Potenciales Align", "Invisalign", lambda r: cell_str(r[1]) if len(r) > 1 else "", 1),
        (F_NORTE, "Potenciales Ormco", "Spark", lambda r: cell_str(r[0]) if r else "", 0),
    ]
    for fname, tab, brand, getname, skip in specs:
        try:
            rows = load_rows(fname, tab)
        except KeyError:
            continue
        for row in rows[skip:]:
            if not row:
                continue
            name = getname(list(row))
            if len(name_tokens(name)) < 2:
                continue
            did = base.match_name(name)
            if did:
                enr.add_brand(did, brand, "%s" % tab)
    # Hoja 10 (B-Learning Bono): ORMCO rows -> Spark
    rows = load_rows(F_BLEARN, "Hoja 10")
    for row in rows:
        cells = list(row)
        if len(cells) < 11:
            continue
        tag = cell_str(cells[10]).upper()
        if "ORMCO" not in tag:
            continue
        name = cell_str(cells[0])
        if len(name_tokens(name)) < 2:
            continue
        did = base.match_name(name)
        if did:
            enr.add_brand(did, "Spark", "Hoja 10")

def src_datos_panel(enr, base, stats):
    """CRM_KS_MX DATOS PANEL: geo prio 1, comercial, accredited_at prio 2, phone/email prio 5."""
    rows = load_rows(F_CRM, "DATOS PANEL")
    hi = find_header(rows, ["Nombre Completo", "Region", "Comercial"])
    if hi is None:
        sys.exit("DATOS PANEL header not found")
    h = rows[hi]
    c_full = col_index(h, "nombre completo")
    c_reg = col_index(h, "region")
    c_com = col_index(h, "comercial")
    c_acc = col_index(h, "fecha de acreditación")
    c_mail = col_index(h, "correo electrónico")
    c_cel = col_index(h, "celular")
    n = 0
    for row in rows[hi + 1:]:
        cells = list(row) + [None] * max(0, len(h) - len(row))
        full = cell_str(cells[c_full])
        if len(name_tokens(full)) < 2:
            continue
        email = clean_email(cells[c_mail]) if c_mail is not None else None
        did, how = base.match(full, email)
        if not did:
            continue
        n += 1
        city, state = split_city_state(cells[c_reg]) if c_reg is not None else (None, None)
        if city or state:
            stats["geo1"].setdefault(did, (city, state, "DATOS PANEL"))
        com = cell_str(cells[c_com]) if c_com is not None else ""
        if com:
            enr.set(did, "comercial_asignado", com, "DATOS PANEL")
        acc = to_date(cells[c_acc]) if c_acc is not None else None
        if acc:
            stats["acc2"].setdefault(did, (acc, "DATOS PANEL"))
        stats["p5"].setdefault(did, {})
        ph = clean_phone(cells[c_cel]) if c_cel is not None else None
        if ph and "phone" not in stats["p5"][did]:
            stats["p5"][did]["phone"] = ph
        if email and "email" not in stats["p5"][did]:
            stats["p5"][did]["email"] = email
    stats["matched_panel"] = n

def src_entregas(enr, base, stats):
    """Entregas 2026->2023: geo prio 2 (latest year wins)."""
    for tab in ("2026", "2025", "2024", "2023"):
        rows = load_rows(F_ENTREGAS, tab)
        hi = find_header(rows, ["ID", "Localidad"])
        if hi is None:
            continue
        h = rows[hi]
        c_dr = col_index(h, "dr/ dra") or col_index(h, "dr/dra")
        c_loc = col_index(h, "localidad")
        for row in rows[hi + 1:]:
            cells = list(row)
            if c_dr is None or c_dr >= len(cells):
                continue
            name = cell_str(cells[c_dr])
            loc = cell_str(cells[c_loc]) if c_loc is not None and c_loc < len(cells) else ""
            if not name or not loc:
                continue
            did = base.match_name(name)
            if not did:
                continue
            city, state = split_city_state(loc)
            if city or state:
                stats["geo2"].setdefault(did, (city, state, "Entregas %s" % tab))

# ================================================================ PAYMENTS

def norm_method(raw):
    m = cell_str(raw)
    if not m:
        return None
    k = strip_accents(m).lower().strip().rstrip(".").strip()
    if k in ("tr", "transferencia"):
        return "TR"
    if k in ("mp", "mercado pago", "mercadopago"):
        return "MP"
    if k.startswith("dep"):
        return "Depósito"
    return m

def parse_payments(base, stats):
    rows = load_rows(F_ADMIN, "Facturación y Cobranzas")
    header = [cell_str(c) for c in rows[0]]
    # locate columns
    def idx_all(pat):
        return [i for i, c in enumerate(header) if re.fullmatch(pat, c.strip(), re.I)]
    id_cols = [i for i, c in enumerate(header) if c.strip().upper() == "ID"]
    c_id = id_cols[1] if len(id_cols) > 1 else id_cols[0]
    c_pac = next(i for i, c in enumerate(header) if c.strip().upper() == "PACIENTE")
    c_prof = next(i for i, c in enumerate(header) if c.strip().upper() == "PROFESIONAL")
    formas = idx_all(r"FORMA DE PAGO")
    fechas_pago = idx_all(r"FECHA PAGO")
    amounts = [i for i, c in enumerate(header) if re.fullmatch(r"\d\s*°\s*PAGO", c.strip(), re.I)]
    nfacs = idx_all(r"N°\s*FAC")
    assert len(formas) == len(fechas_pago) == len(amounts) == 5, (formas, fechas_pago, amounts)

    payments = []
    undated = 0
    prof_match = {}   # raw -> did or None
    for rix, row in enumerate(rows[1:], start=2):  # worksheet row numbers
        cells = list(row) + [None] * max(0, len(header) - len(row))
        case_id = cell_str(cells[c_id])
        if not case_id or case_id.upper() == "ID":
            continue
        prof_raw = cell_str(cells[c_prof])
        if prof_raw not in prof_match:
            prof_match[prof_raw] = base.match_name(prof_raw) if prof_raw else None
        did = prof_match[prof_raw]
        for k in range(5):
            amt = to_amount(cells[amounts[k]])
            pdate = to_date(cells[fechas_pago[k]])
            if amt is None or amt == 0:
                continue
            if not pdate:
                undated += 1
                continue
            method = norm_method(cells[formas[k]])
            nfac = cell_str(cells[nfacs[k]]) if k < len(nfacs) and nfacs[k] < len(cells) else ""
            payments.append({
                "external_key": "adminmx:%d:%d" % (rix, k + 1),
                "doctor_nombre_raw": prof_raw,
                "noloco_id": did,
                "case_external_id": case_id,
                "paciente": cell_str(cells[c_pac]) or None,
                "amount_mxn": round(amt, 2),
                "paid_at": pdate,
                "method": method,
                "notes": ("fac:%s" % nfac) if nfac and nfac != "-" else None,
            })
    stats["payments_undated_amounts"] = undated
    stats["prof_match"] = prof_match
    return payments

# ================================================================ MAIN

def main():
    base = DoctorBase(NOLOCO_JSON)
    enr = Enrichment(base)
    stats = {"geo1": {}, "geo2": {}, "geo3": {}, "geo4": {}, "geo5": {},
             "acc2": {}, "acc3": {}, "p5": {}}

    src_acredi_norte(enr, base, stats)
    src_blearning(enr, base, stats)
    src_curso_tabs(enr, base, stats)
    src_big_bases(enr, base, stats)
    src_competitor_tabs(enr, base, stats)
    src_datos_panel(enr, base, stats)
    src_entregas(enr, base, stats)

    # ---- apply geo by priority: DATOS PANEL > Entregas > Acredi Norte > curso ESTADO > big bases
    for did in base.doctors:
        for level in ("geo1", "geo2", "geo3", "geo4", "geo5"):
            if did in stats[level]:
                city, state, src = stats[level][did]
                if city:
                    enr.set(did, "city", city, src)
                if state:
                    enr.set(did, "state", state, src)
                if enr.rec[did]["city"] or enr.rec[did]["state"]:
                    break
    # ---- accredited_at fallbacks: Acredi Norte already set; then DATOS PANEL; then curso date
    for did in base.doctors:
        if enr.rec[did]["accredited_at"] is None and did in stats["acc2"]:
            enr.set(did, "accredited_at", stats["acc2"][did][0], stats["acc2"][did][1])
        if enr.rec[did]["accredited_at"] is None and did in stats["acc3"]:
            enr.set(did, "accredited_at", stats["acc3"][did][0], stats["acc3"][did][1])
    # ---- phone/email last resort: DATOS PANEL (priority 5)
    for did, vals in stats["p5"].items():
        if "phone" in vals:
            enr.set(did, "phone", vals["phone"], "DATOS PANEL")
        if "email" in vals:
            enr.set(did, "email", vals["email"], "DATOS PANEL")

    # ---- build enrichment.json (only doctors with at least one enriched field)
    out = []
    for did in sorted(base.doctors, key=lambda x: int(x)):
        r = enr.rec[did]
        brands = sorted(enr.brands[did])
        has_any = any(r[f] for f in FIELDS) or brands
        if not has_any:
            continue
        phone = r["phone"]
        city, state = r["city"], r["state"]
        out.append({
            "noloco_id": did,
            "nombre": base.doctors[did]["nombre"],
            "phone": phone,
            "whatsapp": phone,          # same normalized MX mobile number
            "email": r["email"] or base.doctors[did]["email"],
            "city": city,
            "state": state,
            "zona": derive_zona(city, state),
            "accredited_at": r["accredited_at"],
            "competitor_brands": brands,
            "comercial_asignado": r["comercial_asignado"],
            "sources": enr.src[did],
        })

    payments = parse_payments(base, stats)

    os.makedirs(OUT_DIR, exist_ok=True)
    with open(os.path.join(OUT_DIR, "enrichment.json"), "w") as f:
        json.dump(out, f, ensure_ascii=False, indent=1)
    with open(os.path.join(OUT_DIR, "payments.json"), "w") as f:
        json.dump(payments, f, ensure_ascii=False, indent=1)

    write_report(base, enr, out, payments, stats)
    print("enrichment records:", len(out), "| payments:", len(payments))

# ---------------------------------------------------------------- report
EXPECTED_2026 = {  # month -> (sum, count)
    1: (395869, 32), 2: (305319, 26), 3: (310155, 27), 4: (180926, 15),
    5: (243501, 26), 6: (255638, 22), 7: (255465, 23),
}
MONTHS_ES = {1: "Ene", 2: "Feb", 3: "Mar", 4: "Abr", 5: "May", 6: "Jun", 7: "Jul",
             8: "Ago", 9: "Sep", 10: "Oct", 11: "Nov", 12: "Dic"}

def write_report(base, enr, out, payments, stats):
    total = len(base.doctors)
    by_id = {o["noloco_id"]: o for o in out}
    cov = lambda f: sum(1 for o in out if o.get(f))
    n_phone = cov("phone")
    n_city = sum(1 for o in out if o["city"] or o["state"])
    n_acc = cov("accredited_at")
    n_com = cov("comercial_asignado")
    n_brand = sum(1 for o in out if o["competitor_brands"])
    n_email_enriched = sum(1 for o in out if "email" in o["sources"])

    # 2026 monthly validation
    from collections import defaultdict
    monthly = defaultdict(lambda: [0.0, 0])
    for p in payments:
        y, m = int(p["paid_at"][:4]), int(p["paid_at"][5:7])
        if y == 2026:
            monthly[m][0] += p["amount_mxn"]
            monthly[m][1] += 1

    # unmatched ledger doctors
    from collections import Counter
    unmatched = Counter()
    unmatched_amt = Counter()
    for p in payments:
        if p["noloco_id"] is None:
            unmatched[p["doctor_nombre_raw"] or "(vacío)"] += 1
            unmatched_amt[p["doctor_nombre_raw"] or "(vacío)"] += p["amount_mxn"]
    matched_pay = sum(1 for p in payments if p["noloco_id"])

    L = []
    L.append("# Enrichment report — CRM MX import")
    L.append("")
    L.append("Generado: %s por `scripts/parse_enrichment.py`" % datetime.now().strftime("%Y-%m-%d %H:%M"))
    L.append("")
    L.append("## Cobertura (base Noloco: %d doctores; efectivos %d — %d duplicado/s y 1 doctor de prueba)"
             % (total, total - len(base.duplicates) - 1, len(base.duplicates)))
    L.append("")
    L.append("| Campo | Doctores | % de 175 |")
    L.append("|---|---|---|")
    L.append("| Con algún dato enriquecido | %d | %.0f%% |" % (len(out), 100.0 * len(out) / total))
    L.append("| Teléfono (+52 normalizado) | %d | %.0f%% |" % (n_phone, 100.0 * n_phone / total))
    L.append("| Email (de planillas; el resto ya viene de Noloco) | %d | %.0f%% |" % (n_email_enriched, 100.0 * n_email_enriched / total))
    L.append("| Ciudad y/o estado | %d | %.0f%% |" % (n_city, 100.0 * n_city / total))
    L.append("| Zona derivada | %d | %.0f%% |" % (cov("zona"), 100.0 * cov("zona") / total))
    L.append("| Fecha de acreditación | %d | %.0f%% |" % (n_acc, 100.0 * n_acc / total))
    L.append("| Comercial asignado | %d | %.0f%% |" % (n_com, 100.0 * n_com / total))
    L.append("| Marcas competidoras (≥1) | %d | %.0f%% |" % (n_brand, 100.0 * n_brand / total))
    L.append("")
    zonas = Counter(o["zona"] for o in out if o["zona"])
    L.append("Zonas: " + ", ".join("%s %d" % (z, c) for z, c in zonas.most_common()))
    L.append("")
    src_phone = Counter(o["sources"].get("phone") for o in out if o.get("phone") and o["sources"].get("phone"))
    L.append("Fuente del teléfono: " + ", ".join("%s %d" % (s, c) for s, c in src_phone.most_common()))
    L.append("")
    L.append("## Pagos (Administración México → Facturación y Cobranzas)")
    L.append("")
    L.append("- Registros de pago emitidos: **%d** (con fecha e importe)" % len(payments))
    L.append("- Con doctor matcheado a Noloco: %d (%.0f%%)" % (matched_pay, 100.0 * matched_pay / max(1, len(payments))))
    L.append("- Importes sin fecha de pago (omitidos): %d" % stats["payments_undated_amounts"])
    L.append("")
    L.append("### Validación 2026 (esperado vs parseado)")
    L.append("")
    L.append("| Mes | Esperado $ | Parseado $ | Δ$ | Esperado # | Parseado # |")
    L.append("|---|---|---|---|---|---|")
    for m in range(1, 8):
        es, ec = EXPECTED_2026[m]
        ps, pc = monthly.get(m, [0, 0])
        L.append("| %s | %s | %s | %+d | %d | %d |" % (
            MONTHS_ES[m], format(es, ",d"), format(int(round(ps)), ",d"),
            int(round(ps - es)), ec, pc))
    L.append("")
    other_2026 = {m: v for m, v in monthly.items() if m > 7}
    if other_2026:
        L.append("Meses 2026 fuera de la validación: " + ", ".join(
            "%s $%s (%d)" % (MONTHS_ES[m], format(int(v[0]), ",d"), v[1]) for m, v in sorted(other_2026.items())))
        L.append("")
    L.append("### Doctores del ledger sin match en Noloco")
    L.append("")
    if unmatched:
        L.append("| PROFESIONAL (raw) | Pagos | $ total |")
        L.append("|---|---|---|")
        for name, c in unmatched.most_common():
            L.append("| %s | %d | %s |" % (name, c, format(int(unmatched_amt[name]), ",d")))
    else:
        L.append("(ninguno)")
    L.append("")
    L.append("## Anomalías")
    L.append("")
    L.append("- Celdas de teléfono con tipo datetime (formato de la hoja) descartadas: %d" % PHONE_SKIPPED_DATETIMES[0])
    weird_dates = [p for p in payments if p["paid_at"] < "2022-01-01"]
    for p in weird_dates:
        L.append("- Fecha de pago sospechosa `%s` en %s (%s, %s $%s) — probable typo de carga (comparar con FECHA FAC de la fila); se dejó tal cual."
                 % (p["paid_at"], p["external_key"], p["case_external_id"],
                    p["doctor_nombre_raw"], format(int(p["amount_mxn"]), ",d")))
    if base.ambiguous:
        L.append("- Nombres ambiguos (2+ candidatos parejos, se dejaron sin match): %d" % len(base.ambiguous))
        for k, cands in sorted(base.ambiguous.items())[:15]:
            L.append("    - `%s` → %s" % (k, " / ".join(cands)))
    if base.duplicates:
        for dup, canon in sorted(base.duplicates.items()):
            L.append("- Doctor duplicado en Noloco: id %s '%s' (%d casos) tratado como alias de id %s '%s' (%d casos); todo el matching apunta al canónico."
                     % (dup, base.doctors[dup]["nombre"], base.case_counts.get(dup, 0),
                        canon, base.doctors[canon]["nombre"], base.case_counts.get(canon, 0)))
    missing = [base.doctors[d]["nombre"] for d in base.doctors
               if d not in by_id and d not in base.duplicates]
    if missing:
        L.append("- Doctores Noloco sin ningún dato en las planillas (%d): %s" % (len(missing), "; ".join(sorted(missing))))
    L.append("- 'Copia Facturacion' ignorada (copia vieja del ledger).")
    L.append("- Fechas de acreditación de cursos = fecha del curso (aprox., tab del sheet madre).")
    L.append("")
    with open(os.path.join(OUT_DIR, "enrichment_report.md"), "w") as f:
        f.write("\n".join(L))

if __name__ == "__main__":
    main()
