#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""Parser de la caja del consultorio (port de consultorio-gestion/parse_caja.py).

Dos modos:
  --xlsx-a-raw <caja.xlsx> <raw.json>   convierte el export xlsx al MISMO JSON
                                        que devuelve el Apps Script (gas-caja-ar.gs):
                                        {"tabs": {nombre: [[celdas]]}} con fechas
                                        "yyyy-MM-dd" y solo filas 2026.
  --raw-a-movs <raw.json> <movs.json>   parsea ese JSON a los movimientos que
                                        consume import-movimientos-ar.ts.

La clasificación y el orden de recorrido son IDÉNTICOS al parser original:
las external_key del import dependen de eso (contenido + ordinal intra-día).
"""
import re, json, sys, unicodedata, warnings
from datetime import datetime
warnings.filterwarnings('ignore')

TABS = {
    'MONI': 'Mónica González',
    'MARIANA  MATELLI': 'Mariana Matelli',
    'MARIANA KS': 'Mariana Franco',
    'ROCIO 2025': 'Rocío Puig',
    'EUGENIA 2020': 'Eugenia Digiano',
    'CONI 2020': 'Coni',
    'VIRGINIA ': 'Virginia',
}
TAB_SOLICITUD = 'SOLICITUD FACTURAS Y CONSULTAS '

def norm(s):
    s = unicodedata.normalize('NFD', s.lower())
    return ''.join(c for c in s if unicodedata.category(c) != 'Mn')

MESES = ['enero','febrero','marzo','abril','mayo','junio','julio','agosto',
         'septiembre','setiembre','octubre','noviembre','diciembre']

RE_CUOTA = re.compile(r'c(?:uo)?ta\s*\.?\s*(\d+)\s*de\s*(\d+)', re.I)
RE_FECHA = re.compile(r'^\d{4}-\d{2}-\d{2}$')

def classify(paciente, motivo, obs, monto_ars, monto_usd):
    p, m, o = norm(paciente or ''), norm(motivo or ''), norm(obs or '')
    texto = ' '.join([p, m, o])
    neg = (monto_ars is not None and monto_ars < 0) or (monto_usd is not None and monto_usd < 0)
    if neg:
        if 'liquidacion' in texto or 'retiro' in texto or 'rendicion' in texto:
            return 'retiro_liquidacion', None
        if any(k in texto for k in ('lab', 'insumo', 'botones', 'yeso', 'silicona')):
            return 'gasto_tratamiento', None
        return 'gasto_consultorio', None
    if RE_CUOTA.search((motivo or '') + ' ' + (obs or '')) or 'cuota' in texto or 'etapa adicional' in texto:
        return 'cobro', 'Alineadores'
    if 'contenc' in texto:
        return 'cobro', 'Contención'
    if 'consult' in texto or '1era' in texto or '1ra' in texto or 'primera' in texto:
        return 'cobro', 'Consulta'
    if any(mes in texto for mes in MESES) or re.search(r'\bmensualidad', texto):
        return 'cobro', 'Mensualidad'
    return 'cobro', 'Otros'

NOMBRES_DRAS = [
    ('matelli', 'Mariana Matelli'), ('franco', 'Mariana Franco'),
    ('moni', 'Mónica González'), ('monica', 'Mónica González'),
    ('rocio', 'Rocío Puig'), ('eugenia', 'Eugenia Digiano'),
    ('coni', 'Coni'), ('virginia', 'Virginia'),
]

def atribuir_retiro(texto, dra_tab):
    t = norm(texto)
    for clave, nombre in NOMBRES_DRAS:
        if clave in t:
            return nombre, True
    if 'claudia' in t:
        return dra_tab, False
    return dra_tab, False

def parse_tab(rows, tab, dra, year=2026):
    """rows: listas de celdas con datetime ya materializados."""
    movs = []
    for r in rows:
        r = list(r) + [None] * 14
        fecha = fi = None
        for i, c in enumerate(r):
            if isinstance(c, datetime):
                fecha, fi = c, i
                break
        if not fecha or fecha.year != year:
            continue
        paciente, ars, usd, medio, motivo, obs = (
            r[fi+2], r[fi+4], r[fi+5], r[fi+6], r[fi+7], r[fi+8])
        if paciente is None and ars is None and usd is None:
            continue
        def num(x):
            if isinstance(x, (int, float)): return float(x)
            return None
        ars, usd = num(ars), num(usd)
        if ars is None and usd is None:
            continue
        if ars == 0 and (usd is None or usd == 0):
            continue
        tipo, cat = classify(str(paciente or ''), str(motivo or ''), str(obs or ''), ars, usd)
        dra_attr, attr_clara = dra, True
        if tipo == 'retiro_liquidacion':
            dra_attr, attr_clara = atribuir_retiro(
                ' '.join(str(x or '') for x in (paciente, motivo, obs)), dra)
        movs.append({
            'fecha': fecha.strftime('%Y-%m-%d'), 'mes': fecha.month,
            'doctora': dra_attr, 'atribucion_clara': attr_clara, 'tab': tab.strip(),
            'paciente': str(paciente or '').strip(),
            'ars': ars or 0.0, 'usd': usd or 0.0,
            'medio': str(medio or '').strip(), 'motivo': str(motivo or '').strip(),
            'obs': str(obs or '').strip(), 'tipo': tipo, 'categoria': cat,
        })
    return movs

def materializar(rows):
    """celdas "yyyy-MM-dd" (del GAS) → datetime, como las ve openpyxl."""
    out = []
    for r in rows:
        out.append([
            datetime.strptime(c, '%Y-%m-%d') if isinstance(c, str) and RE_FECHA.match(c) else c
            for c in r
        ])
    return out

def xlsx_a_raw(path_xlsx, path_out):
    from openpyxl import load_workbook
    wb = load_workbook(path_xlsx, read_only=True, data_only=True)
    tabs = {}
    for tab in list(TABS) + [TAB_SOLICITUD]:
        if tab not in wb.sheetnames:
            continue
        filas = []
        for r in wb[tab].iter_rows(values_only=True):
            if not any(isinstance(c, datetime) and c.year >= 2026 for c in r):
                continue
            filas.append([
                c.strftime('%Y-%m-%d') if isinstance(c, datetime) else (None if c == '' else c)
                for c in r
            ])
        tabs[tab] = filas
    wb.close()
    json.dump({'tabs': tabs}, open(path_out, 'w'), ensure_ascii=False)
    print(f"raw: {sum(len(v) for v in tabs.values())} filas en {len(tabs)} pestañas")

def raw_a_movs(path_raw, path_out):
    raw = json.load(open(path_raw))
    movs = []
    for tab, dra in TABS.items():
        rows = raw.get('tabs', {}).get(tab)
        if rows is None:
            print(f"AVISO: falta la pestaña '{tab}' en el raw", file=sys.stderr)
            continue
        movs.extend(parse_tab(materializar(rows), tab, dra))
    json.dump(movs, open(path_out, 'w'), ensure_ascii=False, indent=1)
    print(f'{len(movs)} movimientos 2026 extraídos')

if __name__ == '__main__':
    modo = sys.argv[1]
    if modo == '--xlsx-a-raw':
        xlsx_a_raw(sys.argv[2], sys.argv[3])
    elif modo == '--raw-a-movs':
        raw_a_movs(sys.argv[2], sys.argv[3])
    else:
        sys.exit('modo: --xlsx-a-raw <caja.xlsx> <raw.json> | --raw-a-movs <raw.json> <movs.json>')
